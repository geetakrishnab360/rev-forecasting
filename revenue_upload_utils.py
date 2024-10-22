from typing import Tuple, Optional, Union, List, Callable, Any
from pathlib import Path
import pandas as pd
import numpy as np
import re
import difflib
import unidecode
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.utils import get_column_letter
from fuzzywuzzy import fuzz
import snowflake.connector
from io import BytesIO

OA_PROJECT_QUERY = """
WITH project_cte AS
(SELECT 
    name,
    CASE
        WHEN netsuite_subsidiary = 'Blend360 India Limited' THEN 'group1'
        WHEN netsuite_subsidiary = 'Blend360 International Holdings, LLC' THEN 'group1'
        WHEN netsuite_subsidiary = 'Blend360 Switzerland' THEN 'group1'
        WHEN netsuite_subsidiary = 'Blend360, LLC' THEN 'group1'
        WHEN netsuite_subsidiary = 'Engagement Factory AB' THEN 'group4'
        WHEN netsuite_subsidiary = 'Engagement Factory BV Net' THEN 'group4'
        WHEN netsuite_subsidiary = 'Engagement Factory South' THEN 'group4'
        WHEN netsuite_subsidiary = 'Engagement Factory UAE' THEN 'group4'
        WHEN netsuite_subsidiary = 'Engagement Factory UK' THEN 'group4'
        WHEN netsuite_subsidiary = 'Forthpoint' THEN 'group3'
        WHEN netsuite_subsidiary = 'Montevideo Labs' THEN 'group1'
        WHEN netsuite_subsidiary = 'Park Innovate Netherlands' THEN 'group6'
        WHEN netsuite_subsidiary = 'Park Investments' THEN 'group6'
        WHEN netsuite_subsidiary = 'futureproofAI India Private Limited' THEN 'group2'
        WHEN netsuite_subsidiary = 'futureproofAI LTD' THEN 'group2'
        WHEN netsuite_subsidiary = 'futureproofAI, Inc' THEN 'group2'
        WHEN netsuite_subsidiary = 'xElimination Blend International' THEN 'group1'
        WHEN netsuite_subsidiary = 'xElimination Blend360' THEN 'group1'
        WHEN netsuite_subsidiary = 'xElimination futureproofAI Holdings, Inc.' THEN 'group2'
        ELSE netsuite_subsidiary
    END AS group_name,
    netsuite_project_id,
    hubspot_deal_id,
    row_number()  over (partition by name, group_name order by updated desc) as rn
FROM dsx_dashboards.openair_raw.project_raw 
)
SELECT 
    name,
    group_name,
    netsuite_project_id,
    hubspot_deal_id
FROM project_cte
WHERE rn = 1
"""

DATA_PATH = Path("./")

CLIENT_RENAME_MAPPING = {
    "AT&T": "ATT",
    "AmEx": "American Express Company",
    "Amex": "American Express Company",
    "American Express": "American Express Company",
    "Amperity EF SA": "Amperity",
    "Blackstone Inc.": "The Blackstone Group",
    "CompNova, LLC": "ATT",
    "Cornerstone Building Brands": "Clayton, Dubilier & Rice",
    "FEDERACIO FARMACEUTICA, S.C.C.L": "Fedefarma",
    "FEDERACIAfaEURoe FARMACAfE+UTICA, S.C.C.L": "Fedefarma",
    "Flatiron Health, Inc.": "Flatiron Health",
    "Globalization Partners Inc.": "Globalization Partners",
    "Huntsworth Health Limited (Inizio)": "Inizio Healthcare Limited",
    "Multi-Color Corporation (MCC)": "Clayton, Dubilier & Rice",
    "OneLook Inc. D/B/A Lily AI": "lily.ai",
    "Pragmatic - LinkedIn": "Pragmatic Communications",
    "Tech Mahindra Americas": "ATT",
    "Universal Parks & Resorts": "UDX",
    "Wyndham": "Wyndham Hotels and Resorts, LLC",
    "Visa (BTS)": "Visa",
    "Citi": "citigroup",
    "Blackstone": "The Blackstone Group",
    "IAC": "IAC/InterActiveCorp",
    "Legrand": "Legrand Cablofil Inc",
    "Wyndham Hotels and Resorts": "Wyndham Hotels and Resorts, LLC",
    "The William Carter Company": "Carter's",
    "Stagwell Marketing Cloud LLC": "Stagwellglobal",
    "Sincera Technologies LLC": "Sincera",
    "Session M Inc.": "MasterCard",
    "LinkedIn Corporation": "LinkedIn",
    "Ocean Spray": "OceanSpray",
    "Pixability Inc.": "Pixability",
    "Roku Inc.": "Roku",
    "BT": "British Telecom",
    "Blend360": "Blend360: DSX Internal",
    "Newfold Digital": "Newfold",
    "Templeton": "Franklin Templeton",
}

DEPT_LOOKUP = {
    "Interco - Data Science Revenue": "Data Science",
    "Interco - Insourcing Services Revenue": "Insourcing Services",
    "Interco - FPAI Revenue": "FPAI -General",
    "Interco - Data Engineering Revenue": "Data Engineering",
    "Interco - Business Intelligence Revenue": "Business Intelligence",
    "Interco - AI Consulting Revenue": "AI Consulting",
    "Interco - Solutions Revenue": "Solutions",
    "Interco - CX Revenue": "CX",
    "Interco - Intercompany MarkUp Rev": "Intercompany",
    "Revenue from CX projects": "CX",
}

SUBSIDIARY_GROUPS = {
    "Blend360 India Limited": "group1",
    "Blend360 Switzerland": "group1",
    "Blend360, LLC": "group1",
    "Engagement Factory AB": "group4",
    "Engagement Factory BV Net": "group4",
    "Engagement Factory South": "group4",
    "Engagement Factory UAE": "group4",
    "Engagement Factory UK": "group4",
    "Forthpoint": "group3",
    "Montevideo Labs": "group1",
    "Park Innovate Netherlands": "group6",
    "Park Investments": "group6",
    "futureproofAI India Private Limited": "group2",
    "futureproofAI LTD": "group2",
    "futureproofAI, Inc": "group2",
    "xElimination Blend International": "group1",
    "xElimination Blend360": "group1",
    'Blend360 International Holdings, LLC' : 'group1',
    "xElimination futureproofAI Holdings, Inc.": "group2",
}

ADDITIONAL_CLIENTS = [
                    "BT",
                    "Amex",
                    "Legrand",
                    "Wyndham",
                    "British Telecom",
                    "Ferrero",
                    "Templeton",
                    "FEDERACIAfaEURoe FARMACAfE+UTICA, S.C.C.L",
                ]

def fetch_data(sql_query):
    try:
        conn = snowflake.connector.connect(
            user="forecasting_app_service_account",
            account="c2gpartners.us-east-1",
            password="B!_ForecastApp2024",
            warehouse="POWERHOUSE",
        )
        cur = conn.cursor()
        cur.execute(sql_query)
        data = pd.DataFrame(
            cur.fetchall(), columns=[desc[0] for desc in cur.description]
        )
        conn.close()
        return data
    except Exception as e:
        print(e)
        pass
    
def read_file(
    file_name: str, sheet_name: str = "Original", header_row: int = 6
) -> pd.DataFrame:
    """Reads an Excel file and returns a Pandas DataFrame.

    Args:
        file_name (str): The path to the Excel file.
        sheet_name (str, optional): The name of the sheet to read.
                                    Defaults to 'Original'.
        header_row (int, optional): The row number to use as the header.
                                    Defaults to 6.

    Returns:
        pd.DataFrame: The DataFrame containing the data from the Excel file.
    """
    df = pd.read_excel(file_name, sheet_name=sheet_name, header=header_row)
    return df


def strip_spaces(func: Callable) -> Callable:
    def __wrapper__(*args: List[Any]) -> List[Optional[str]]:
        res = func(*args)
        return [x.strip() if isinstance(x, str) else x for x in res]

    return __wrapper__


@strip_spaces
def extract_client_and_project_from_entity(
    entity_line: Union[str, float], client_list: List[str]
) -> List[Optional[str]]:

    if not isinstance(entity_line, str):
        return [np.nan, np.nan]
    

    if "ICC" in entity_line:
        pieces = entity_line.split(":")
        client_match = extract_matching_options(pieces[0], client_list)
        if client_match:
            return [client_match, ":".join(pieces[1:])]
        else:
            if "FEDERACI" in pieces[0]:
                return ["Fedefarma", pieces[1]]
            if len(pieces) == 3:
                return pieces[1:]
            else:
                if "-" in pieces[1]:
                    return [pieces[1].split("-", maxsplit=1)[0], pieces[1]]
                return [pieces[1], pieces[1]]

    else:
        if ":" not in entity_line:
            return [entity_line, np.nan]
        else:
            return entity_line.split(":", maxsplit=1)


def preprocess(
    input_df: pd.DataFrame, client_list: List[str]
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Preprocesses the input DataFrame by dropping rows with missing values.

    Args:
        input_df (pd.DataFrame): The DataFrame to preprocess.

    Returns:
        Tuple[pd.DataFrame, pd.DataFrame]: The preprocessed DataFrame.
    """
    df = input_df.copy()
    # Replace - No Entity - with NaN
    df["Entity (Line)"] = df["Entity (Line)"].replace("- No Entity -", np.nan)
    # Drop rows with missing values in Type
    df = df[~df["Type"].isna()].copy()

    df["department_revised"] = df["Department: Name"].apply(revise_department)

    # Drop rows with Department: Name is Intercompany
    intercompany_txn = df[df["Department: Name"] == "Intercompany"].copy()
    df = df[~(df["Department: Name"] == "Intercompany")].copy()
    # Client name extraction will be done from Entity (Line), when missings
    # fall back to ICC projects
    df["entity_line_coalesced"] = (
        df[["Entity (Line)", "ICC Projects"]]
        .bfill(axis=1)
        .infer_objects(copy=False)
        .iloc[:, 0]
        .copy()
    )

    df[["client", "project"]] = df.entity_line_coalesced.apply(
        extract_client_and_project_from_entity, args=(client_list,)
    ).tolist()
    df.project = df.project.str.replace(
        r"^\d+(?!:)", "", regex=True
    ).str.strip()
    df.client = df.client.str.replace(r"^(\d)+", "", regex=True).str.strip()
    df["client_from_description"] = df["Description"].apply(
        lambda x: extract_matching_options(x, client_list)
    )
    df["client"] = df["client"].fillna(df["client_from_description"])

    clients = df.client.dropna().unique()
    mapping = {
        client: extract_matching_options(client, client_list) or client
        for client in clients
        if client.lower() not in map(lambda x: x.lower(), client_list)
    }
    df.client = df.client.apply(
        lambda x: unidecode.unidecode(x) if isinstance(x, str) else x
    )
    df.client = df.client.replace(mapping).replace(CLIENT_RENAME_MAPPING)
    df["sub_group"] = df["Subsidiary: Name"].map(SUBSIDIARY_GROUPS)
    for col in [
        "Entity (Line): Internal ID",
        "Entity (Line): HubSpot Deal ID",
    ]:
        df[col] = df.groupby(["client", "project", "sub_group"])[
            col
        ].transform(lambda x: x.ffill().bfill())
    df["Match score"] = df[
        ["entity_line_coalesced", "Description", "client"]
    ].apply(
        lambda row: (
            fuzz.partial_ratio(
                (
                    row["entity_line_coalesced"]
                    if not pd.isnull(row["entity_line_coalesced"])
                    else (
                        row["Description"]
                        if not pd.isnull(row["Description"])
                        else ""
                    )
                ),
                row["client"],
            )
            if isinstance(row["client"], str)
            else 0
        ),
        axis=1,
    )
    project_df = fetch_data(OA_PROJECT_QUERY)
    df = df.merge(
        project_df,
        left_on=["project", "sub_group"],
        right_on=["NAME", "GROUP_NAME"],
        how="left",
    ).copy()

    df["Entity (Line): Internal ID"].fillna(
        df["NETSUITE_PROJECT_ID"], inplace=True
    )
    df["Entity (Line): HubSpot Deal ID"].fillna(
        df["HUBSPOT_DEAL_ID"], inplace=True
    )

    df.project = df.project.fillna("NA")
    df.client = df.client.fillna("NA")
    intercompany_txn["client"] = "ZZ_Elim"
    return (df, intercompany_txn)


def drop_clients_from_list(clients: pd.Series) -> pd.Series:
    clients = clients[
        (~clients.str.startswith("ICC"))
        & (~clients.str.startswith("CC"))
        & (~clients.str.contains("Blend"))
    ]
    return clients


def revise_department(department):
    if not isinstance(department, str):
        return
    if department == "ForthPoint":
        return "ForthPoint"
    elif department == "Intercompany":
        return DEPT_LOOKUP.get(department)
    else:
        return department


def read_client_list(file_name: str = "Clients619.xlsb") -> List[str]:
    client_name_df = pd.read_excel(file_name)
    client_name_df.info()
    client_list = (
        client_name_df[["Name", "Client Nickname"]]
        .stack()
        .dropna()
        .drop_duplicates()
        .pipe(drop_clients_from_list)
    ).tolist()
    return client_list


def match_score(text: str, option: str) -> Tuple[str, float]:
    text = unidecode.unidecode(text)
    option = unidecode.unidecode(option)
    text = text.lower()
    option_ = option.lower()
    option_ = re.sub(r"^(\d+|the|h\.)", "", option_).replace(".", " ")
    text = text.replace(".", " ")
    matcher = difflib.SequenceMatcher(None, text, option_)
    match = matcher.find_longest_match(0, len(text), 0, len(option_))
    longest_match = text[match.a : match.a + match.size].strip()
    # if len(longest_match) == len(option)
    pattern = rf"\b{re.escape(longest_match)}\b".replace("\ ", " ")
    full_word_match = re.match(pattern, option_)
    full_word_match_input = re.search(
        pattern, text.replace("_", " ").replace("-", " ")
    )
    if full_word_match and full_word_match_input:
        score = round(100 * len(longest_match) / len(option_), 2)
    else:
        score = 0
    return option, score


def extract_matching_options(
    text: Optional[str],
    options: list[str],
    threshold: int = 50,
) -> Optional[str]:
    if not isinstance(text, str):
        return
    if text.startswith('x'):
        return ("ZZ_Elim", 0)
    if "point" in [o.lower() for o in options]:
        match = re.search(
            r"(?<!forth)point", text.replace(" ", ""), re.IGNORECASE
        )

        if not match:
            options = [o for o in options if o.lower() != "point"]

    matches = [match_score(text, option) for option in options]

    matches = list(filter(lambda x: x[1] >= threshold, matches))
    if len(matches) > 0:
        matches = sorted(
            matches,
            key=lambda x: x[1] * 100 + len(x[0]),
            reverse=True,
        )
        return matches[0][0]


def create_and_format_worksheet(
    wb: Workbook,
    sheet_name: str,
    data: pd.DataFrame,
    header_font: Font,
    header_fill: PatternFill,
    numerical_cols: list[str],
    null_fill: PatternFill,
    null_fill_col: list = [],
):
    """Creates a new worksheet in the workbook, writes data, and applies formatting."""

    ws = wb.create_sheet(sheet_name)

    # Write headers
    for col_num, value in enumerate(data.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = value
        cell.font = header_font
        cell.fill = header_fill

    # Write data
    for row_num, row in enumerate(data.itertuples(), 2):
        for col_num, value in enumerate(row[1:], 1):  # Skip index
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            col_name = data.columns[col_num - 1]
            if col_name in numerical_cols:
                cell.number_format = "#,##0"
            if col_name == null_fill_col and value == "NA":
                cell.fill = null_fill

    # Auto-fit column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            cell.alignment = Alignment(wrap_text=True)
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (
            max_length * 1.1 if max_length < 20 else max_length * 0.5
        )
        ws.column_dimensions[get_column_letter(column[0].column)].width = (
            adjusted_width
        )


def insert_dict_as_columns(
    data_dict: dict[str, str],
    sheet_name: str,
    wb: Workbook,
    header_names: list[str],
    header_font: Font,
    header_fill: PatternFill,
):
    """
    Inserts a dictionary into an Excel worksheet as two columns (keys and values).

    Args:
        data_dict (dict): The dictionary to insert.
        sheet_name (str): The name of the sheet to insert into.
        wb (openpyxl.Workbook): The openpyxl workbook object.
        header_names (list): A list of two strings representing the header names
                             for the key and value columns respectively.
        header_font (openpyxl.styles.Font): Font style for the headers.
        header_fill (openpyxl.styles.PatternFill): Fill style for the headers.
    """

    # Create the worksheet if it doesn't exist
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]

    # Write headers
    ws.cell(row=1, column=1).value = header_names[0]
    ws.cell(row=1, column=2).value = header_names[1]
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=2).font = header_font
    ws.cell(row=1, column=2).fill = header_fill

    # Write dictionary data
    for row_num, (key, value) in enumerate(data_dict.items(), start=2):
        ws.cell(row=row_num, column=1).value = key
        ws.cell(row=row_num, column=2).value = value

    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(column[0].column)].width = (
            adjusted_width
        )


def color_match_scores(ws, match_score_column, start_row=2):
    """
    Colors the match_score column in the worksheet based on the score values.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        match_score_column (int or str): The column index (int) or letter (str)
                                         containing the match_score values.
        start_row (int, optional): The row number to start coloring from.
                                    Defaults to 2 (assuming header in row 1).
    """

    # Define subtle color gradients (RGB values)
    green_100 = Color(rgb="E0F2DE")  # Most green
    green_80 = Color(rgb="D8F0D1")
    green_50 = Color(rgb="D1ECD8")
    red_40 = Color(rgb="FFEBEB")
    red_0 = Color(rgb="FFDCDC")

    # Convert column index to letter if needed
    if isinstance(match_score_column, int):
        match_score_column = get_column_letter(match_score_column)

    for row in range(start_row, ws.max_row + 1):
        cell = ws[f"{match_score_column}{row}"]
        # print(match_score_column)
        score = cell.value
        if pd.isnull(score):
            continue
        elif score >= 100:
            cell.fill = PatternFill(
                start_color=green_100, end_color=green_100, fill_type="solid"
            )
        elif 80 <= score < 100:
            cell.fill = PatternFill(
                start_color=green_80, end_color=green_80, fill_type="solid"
            )
        elif 50 <= score < 80:
            cell.fill = PatternFill(
                start_color=green_50, end_color=green_50, fill_type="solid"
            )
        elif 40 <= score < 50:
            cell.fill = PatternFill(
                start_color=red_40, end_color=red_40, fill_type="solid"
            )
        elif score < 40:
            cell.fill = PatternFill(
                start_color=red_0, end_color=red_0, fill_type="solid"
            )


def export_txn_df_to_excel(
    act_df: pd.DataFrame,
    mod_df: pd.DataFrame,
    client_list: list[str],
    numerical_cols: list[str],
    null_fill_col: list[str],
    filename: str = "netsuite_revenue.xlsx",
):
    """
    Exports the txn_df DataFrame to an Excel file with formatting, enhancements, and pivot tables.

    Args:
        txn_df: The pandas DataFrame to export.
        numerical_cols: A list of column names to format as numbers with comma separation and no decimals.
        null_fill_col: The name of the column to apply light red fill to cells with null values.
        filename: The name of the Excel file to create.
    """

    wb = Workbook()

    del wb["Sheet"]

    # Define formatting styles
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="00A0C6F5", end_color="00A0C6F5", fill_type="solid"
    )
    null_fill = PatternFill(
        start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
    )

    create_and_format_worksheet(
        wb,
        "Original",
        act_df,
        header_font,
        header_fill,
        [],
        null_fill,
        [],
    )

    create_and_format_worksheet(
        wb,
        "Modified",
        mod_df,
        header_font,
        header_fill,
        numerical_cols,
        null_fill,
        null_fill_col,
    )

    color_match_scores(
        wb["Modified"], list(mod_df.columns).index("Match score") + 1
    )

    # --- Create Client Wise Pivot Table ---
    client_pivot_data = pd.pivot_table(
        mod_df, values="Amount", index=["client"], aggfunc=sum
    )
    client_pivot_data.loc["Total"] = client_pivot_data["Amount"].sum()
    create_and_format_worksheet(
        wb,
        "Client Wise Pivot",
        client_pivot_data.reset_index(),
        header_font,
        header_fill,
        numerical_cols,
        null_fill,
        null_fill_col,
    )

    # --- Create Project Wise Pivot Table ---
    project_pivot_data = pd.pivot_table(
        mod_df, values="Amount", index=["client", "project"], aggfunc=sum
    )
    total_project_amount = project_pivot_data["Amount"].sum()
    project_pivot_data.loc[("Total", "")] = total_project_amount
    create_and_format_worksheet(
        wb,
        "Project Wise Pivot",
        project_pivot_data.reset_index(),
        header_font,
        header_fill,
        numerical_cols,
        null_fill,
        null_fill_col,
    )

    insert_dict_as_columns(
        DEPT_LOOKUP,
        "Department Lookup",
        wb,
        ["Account", "Department"],
        header_font,
        header_fill,
    )
    insert_dict_as_columns(
        CLIENT_RENAME_MAPPING,
        "Client Rename Mapping",
        wb,
        ["Client", "Renamed Client"],
        header_font,
        header_fill,
    )

    create_and_format_worksheet(
        wb,
        "Client List",
        pd.DataFrame(client_list, columns=["Clients"]),
        header_font,
        header_fill,
        [],
        null_fill,
        [],
    )

    # Save the workbook
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
    
def update_client_list(client_list: list[str]) -> list[str]:
    if "Forthpoint" in client_list:
        client_list.remove("Forthpoint")
    if "Forth Point" in client_list:
        client_list.remove("Forth point")
    if "Forthpoint (Internal)" in client_list:
        client_list.remove("Forthpoint (Internal)")
    if "Point" in client_list:
        client_list.remove("Point")
        
    client_list.extend(ADDITIONAL_CLIENTS)
    return [unidecode.unidecode(c) for c in client_list]